import io

import openpyxl
import pytest

from app.exceptions.quiz_exceptions import InvalidQuizFileException
from app.utils.excel_importer import ExcelImporter

HEADER = [
    "quiz_title", "quiz_description", "quiz_frequency",
    "question_title", "question_type",
    "answer_1_text", "answer_1_correct",
    "answer_2_text", "answer_2_correct",
    "answer_3_text", "answer_3_correct",
    "answer_4_text", "answer_4_correct",
]


def make_workbook_bytes(rows: list[list]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(HEADER)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def make_quiz_row(**overrides) -> list:
    defaults = dict(
        quiz_title="Python Basics",
        quiz_description="A quiz about Python",
        quiz_frequency=7,
        question_title="What is a list?",
        question_type="single answer",
        answer_1_text="A mutable sequence",
        answer_1_correct=True,
        answer_2_text="A number",
        answer_2_correct=False,
        answer_3_text=None,
        answer_3_correct=None,
        answer_4_text=None,
        answer_4_correct=None,
    )
    defaults.update(overrides)
    return [defaults[col] for col in HEADER]


def test_parse_single_quiz_with_two_questions():
    file_bytes = make_workbook_bytes([
        make_quiz_row(),
        make_quiz_row(question_title="What is a tuple?"),
    ])

    quizzes = ExcelImporter.parse_quizzes(file_bytes)

    assert len(quizzes) == 1
    quiz = quizzes[0]
    assert quiz.title == "Python Basics"
    assert quiz.description == "A quiz about Python"
    assert quiz.frequency == 7
    assert len(quiz.questions) == 2
    assert quiz.questions[0].title == "What is a list?"
    assert len(quiz.questions[0].answers) == 2
    assert quiz.questions[0].answers[0].is_correct is True


def test_parse_groups_rows_with_same_quiz_title():
    file_bytes = make_workbook_bytes([
        make_quiz_row(question_title="Q1"),
        make_quiz_row(question_title="Q2"),
    ])

    quizzes = ExcelImporter.parse_quizzes(file_bytes)

    assert len(quizzes) == 1
    assert len(quizzes[0].questions) == 2
    assert [q.title for q in quizzes[0].questions] == ["Q1", "Q2"]


def test_parse_separates_different_quiz_titles():
    file_bytes = make_workbook_bytes([
        make_quiz_row(quiz_title="Quiz A", question_title="A-Q1"),
        make_quiz_row(quiz_title="Quiz A", question_title="A-Q2"),
        make_quiz_row(quiz_title="Quiz B", question_title="B-Q1"),
        make_quiz_row(quiz_title="Quiz B", question_title="B-Q2"),
    ])

    quizzes = ExcelImporter.parse_quizzes(file_bytes)

    assert {q.title for q in quizzes} == {"Quiz A", "Quiz B"}


def test_parse_reads_up_to_four_answers():
    file_bytes = make_workbook_bytes([
        make_quiz_row(
            question_type="multiple answer",
            answer_3_text="A boolean", answer_3_correct=False,
            answer_4_text="A string", answer_4_correct=True,
        ),
        make_quiz_row(question_title="What is a tuple?"),
    ])

    quizzes = ExcelImporter.parse_quizzes(file_bytes)

    assert len(quizzes[0].questions[0].answers) == 4


def test_parse_skips_blank_rows():
    file_bytes = make_workbook_bytes([
        make_quiz_row(),
        make_quiz_row(question_title="What is a tuple?"),
        [None] * len(HEADER),
    ])

    quizzes = ExcelImporter.parse_quizzes(file_bytes)

    assert len(quizzes) == 1
    assert len(quizzes[0].questions) == 2


def test_parse_raises_on_missing_columns():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["quiz_title", "question_title"])
    sheet.append(["Python Basics", "What is a list?"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(buffer.getvalue())


def test_parse_raises_on_missing_quiz_title():
    file_bytes = make_workbook_bytes([make_quiz_row(quiz_title=None)])

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(file_bytes)


def test_parse_raises_on_invalid_question_type():
    file_bytes = make_workbook_bytes([make_quiz_row(question_type="yes/no")])

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(file_bytes)


def test_parse_raises_when_fewer_than_two_answers():
    file_bytes = make_workbook_bytes([make_quiz_row(answer_2_text=None, answer_2_correct=None)])

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(file_bytes)


def test_parse_raises_on_empty_file():
    workbook = openpyxl.Workbook()
    buffer = io.BytesIO()
    workbook.save(buffer)

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(buffer.getvalue())


def test_parse_raises_on_invalid_workbook_bytes():
    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(b"not a real xlsx file")


def test_parse_raises_on_invalid_frequency():
    file_bytes = make_workbook_bytes([make_quiz_row(quiz_frequency=None)])

    with pytest.raises(InvalidQuizFileException):
        ExcelImporter.parse_quizzes(file_bytes)
